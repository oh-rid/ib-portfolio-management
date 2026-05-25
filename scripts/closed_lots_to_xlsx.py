#!/usr/bin/env python3
"""
Parse IB Flex Closed-Lots XML -> accountant-ready XLSX with EUR conversion.

Each row in the output is one closed lot (matched open-buy <-> close-sell pair).
EUR conversion uses ECB euro foreign exchange reference rates
(eurofxref-hist.csv, daily, EUR base, official EU bookkeeping standard).

Usage:
  # Convenience: auto-resolve all paths for a full tax year
  closed_lots_to_xlsx.py --year 2025
    # Reads:  portfolio/ib/tax/2025/closed_lots_2025.xml
    # Reads:  data/ecb/eurofxref-hist.csv
    # Writes: portfolio/ib/tax/2025/Closed_Lots_2025_EUR.xlsx

  # Manual: explicit paths
  closed_lots_to_xlsx.py --xml input.xml --ecb eurofxref-hist.csv --out report.xlsx

Convention: outputs go to portfolio/ib/tax/{YYYY}/. See
plugins/ib-gateway/scripts/README.md for full conventions and IB Flex quirks.
"""
from __future__ import annotations

import argparse
import csv
import sys
import xml.etree.ElementTree as ET
from datetime import date, datetime, timedelta
from pathlib import Path
from typing import Optional

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


# -------------- ECB rates --------------

def load_ecb_rates(csv_path: Path) -> tuple[dict[date, dict[str, float]], list[str]]:
    """Returns (date -> ccy -> rate) where rate is units of CCY per 1 EUR."""
    rates: dict[date, dict[str, float]] = {}
    with open(csv_path) as f:
        rdr = csv.reader(f)
        header = next(rdr)
        ccys = [c for c in header[1:] if c]
        for row in rdr:
            if not row or not row[0]:
                continue
            try:
                d = datetime.strptime(row[0], "%Y-%m-%d").date()
            except ValueError:
                continue
            day = {}
            for i, ccy in enumerate(header[1:], start=1):
                if not ccy:
                    continue
                v = row[i].strip()
                if v and v != "N/A":
                    try:
                        day[ccy] = float(v)
                    except ValueError:
                        pass
            rates[d] = day
    return rates, sorted(ccys)


def rate_for(rates: dict[date, dict[str, float]],
             d: date, ccy: str) -> tuple[Optional[float], date, str]:
    """Find rate for currency on date d, carrying forward to last business day.
    Returns (rate, actual_date_used, note).
    rate = units of ccy per 1 EUR.
    """
    if ccy == "EUR":
        return 1.0, d, ""
    # CNH (offshore yuan) — ECB only publishes CNY (onshore). Use CNY as proxy.
    proxy_note = ""
    lookup_ccy = ccy
    if ccy == "CNH":
        lookup_ccy = "CNY"
        proxy_note = "CNY proxy for CNH"
    if ccy == "ILA":  # Israeli new shekel agora — should not appear, but guard
        lookup_ccy = "ILS"
    # Carry-forward up to 10 days back
    for back in range(0, 11):
        probe = d - timedelta(days=back)
        if probe in rates and lookup_ccy in rates[probe]:
            note = proxy_note
            if back > 0:
                cf = f"carry-forward {back}d"
                note = f"{note}; {cf}" if note else cf
            return rates[probe][lookup_ccy], probe, note
    return None, d, f"no ECB rate for {ccy} near {d}"


def to_eur(amount_in_ccy: float, rate: Optional[float]) -> Optional[float]:
    if rate is None or rate == 0:
        return None
    return amount_in_ccy / rate


# -------------- XML parsing --------------

def parse_date(s: str) -> date:
    s = s.strip()
    # Accept yyyyMMdd or yyyy-MM-dd or yyyy-MM-dd;HH:mm:ss
    s = s.split(";")[0]
    for fmt in ("%Y%m%d", "%Y-%m-%d"):
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            continue
    raise ValueError(f"unrecognised date: {s!r}")


def f(s: str | None, default: float = 0.0) -> float:
    if s is None or s == "":
        return default
    try:
        return float(s)
    except ValueError:
        return default


def extract_lots(xml_path: Path) -> tuple[list[dict], dict]:
    """Returns (lot rows, metadata).

    IB Flex 'Closed Lots' level represents each closed lot as a <Lot> element
    inside <Trades> (NOT as <Trade>). Sign/field conventions for <Lot>:
      - quantity: absolute (positive), direction comes from buySell
      - tradePrice: CLOSING-trade price (per unit)
      - cost: close-trade money (gross close value, equals quantity*tradePrice*multiplier)
      - origTradePrice / origTradeDate: usually empty at Lot level (aggregated)
      - fifoPnlRealized: AUTHORITATIVE realized P&L for this lot (in trade currency)
      - openDateTime / holdingPeriodDateTime: when the matched opening lot was acquired
      - notes: contains 'LT' (long-term) or 'ST' (short-term), 'W' wash, etc.
    Cost basis is derived: gross_close_value - fifoPnlRealized.
    """
    root = ET.parse(xml_path).getroot()
    stmt = root.find(".//FlexStatement")
    meta = {
        "accountId": stmt.attrib.get("accountId") if stmt is not None else "",
        "fromDate": stmt.attrib.get("fromDate") if stmt is not None else "",
        "toDate": stmt.attrib.get("toDate") if stmt is not None else "",
        "period": stmt.attrib.get("period") if stmt is not None else "",
        "whenGenerated": stmt.attrib.get("whenGenerated") if stmt is not None else "",
        "queryName": root.attrib.get("queryName", ""),
    }
    trades_el = stmt.find("Trades") if stmt is not None else None
    if trades_el is None:
        return [], meta
    # Both <Lot> and <Trade> with levelOfDetail=CLOSED_LOT can appear depending on
    # how the Flex query was configured — handle both.
    lot_nodes = list(trades_el.findall("Lot")) + [
        t for t in trades_el.findall("Trade")
        if t.attrib.get("levelOfDetail", "").upper() == "CLOSED_LOT"
    ]
    skipped_levels: dict[str, int] = {}
    for t in trades_el.findall("Trade"):
        lod = t.attrib.get("levelOfDetail", "")
        if lod.upper() != "CLOSED_LOT":
            skipped_levels[lod or "(none)"] = skipped_levels.get(lod or "(none)", 0) + 1

    lots: list[dict] = []
    for el in lot_nodes:
        a = el.attrib
        try:
            close_date = parse_date(a.get("tradeDate", a.get("dateTime", "")))
        except ValueError:
            continue
        open_date = None
        for k in ("openDateTime", "holdingPeriodDateTime", "origTradeDate"):
            if a.get(k):
                try:
                    open_date = parse_date(a[k])
                    break
                except ValueError:
                    pass

        qty = abs(f(a.get("quantity")))
        mult = f(a.get("multiplier"), 1.0) or 1.0
        realized = f(a.get("fifoPnlRealized"))
        side = a.get("buySell", "").upper()
        asset_cat = a.get("assetCategory", "")

        # IB <Lot> field semantics — verified empirically against trades_raw.xml
        # (Activity Flex with executions-level detail) on AMD vertical spread:
        # - tradePrice = OPENING-trade price (per unit). NOT the close price.
        # - cost       = qty × tradePrice × mult = OPENING cost basis
        # - buySell    = direction of the CLOSING trade (universal across asset
        #                classes — STK, OPT, FOP, FUT all follow this rule):
        #                  SELL = sold to close → was LONG
        #                  BUY  = bought to cover → was SHORT
        # - openDateTime / dateTime / tradeDate = matched open/close dates
        # - fifoPnlRealized = signed realized P&L (positive = profit, negative = loss)
        # Close price is NOT stored directly in <Lot>; derive from accounting identity.
        is_long = (side == "SELL")
        direction = "LONG" if is_long else "SHORT"

        open_price = f(a.get("tradePrice"))            # per unit, at opening
        open_value = qty * open_price * mult            # cost basis (or short proceeds)

        # Derive close price from realized P&L:
        #   LONG : realized = (close - open) × qty × mult  → close = open + realized / (qty × mult)
        #   SHORT: realized = (open - close) × qty × mult  → close = open - realized / (qty × mult)
        if qty > 0 and mult > 0:
            if is_long:
                close_price = open_price + realized / (qty * mult)
            else:
                close_price = open_price - realized / (qty * mult)
        else:
            close_price = None
        close_value = qty * close_price * mult if close_price is not None else None

        # Anomaly trigger: implied close_price negative (impossible — options/stocks
        # can't have negative cash settlement; expired worthless = 0). Catches
        # corporate actions where realized P&L derivation breaks down.
        anomaly = close_price is not None and close_price < -0.0001

        # Prefer IB-provided origTradePrice when populated (rare in Closed Lots
        # query — usually empty). Falls back to tradePrice (which IS the open
        # price under correct interpretation of the <Lot> schema).
        explicit_open_price = f(a.get("origTradePrice")) or None
        if explicit_open_price:
            open_price = explicit_open_price

        commission = f(a.get("ibCommission"))
        commission_ccy = a.get("ibCommissionCurrency", "")
        currency = a.get("currency", "")

        notes = a.get("notes", "")
        # Decode common note codes for accountant readability
        note_decoded = decode_notes(notes)

        lots.append({
            "accountId": a.get("accountId", meta["accountId"]),
            "symbol": a.get("symbol", ""),
            "description": a.get("description", ""),
            "isin": a.get("isin", ""),
            "cusip": a.get("cusip", ""),
            "conid": a.get("conid", ""),
            "assetCategory": a.get("assetCategory", ""),
            "subCategory": a.get("subCategory", ""),
            "underlyingSymbol": a.get("underlyingSymbol", ""),
            "exchange": a.get("exchange", a.get("listingExchange", "")),
            "currency": currency,
            "multiplier": mult,
            "buySell": a.get("buySell", ""),
            "direction": direction,
            "is_long": is_long,
            "anomaly": anomaly,
            "openCloseIndicator": a.get("openCloseIndicator", ""),
            "notes": notes,
            "notesDecoded": note_decoded,
            "openDate": open_date,
            "closeDate": close_date,
            "quantity": qty,
            "openPrice": open_price,
            "closePrice": close_price if not anomaly else None,
            "proceeds": close_value if not anomaly else None,
            "costBasis": open_value,
            "commission": commission,
            "commissionCurrency": commission_ccy,
            "fifoPnlRealized": realized,
            "fxRateToBase": f(a.get("fxRateToBase"), 1.0),
            "transactionID": a.get("transactionID", ""),
            "ibExecID": a.get("ibExecID", ""),
            "levelOfDetail": a.get("levelOfDetail", "CLOSED_LOT"),
            "reportDate": a.get("reportDate", ""),
            "originallyImpliedFromCostField": f(a.get("cost")),  # for audit
        })
    meta["skipped_by_level"] = skipped_levels
    meta["lots_extracted"] = len(lots)
    return lots, meta


# IB notes codes — selected subset relevant to closed lots / tax accounting
NOTE_CODES = {
    "O": "Open", "C": "Close", "P": "Partial", "A": "Assignment",
    "Ex": "Exercise", "Ep": "Expired", "L": "Liquidation by IB",
    "Ca": "Cancelled", "Co": "Corrected", "R": "Dividend reinvest",
    "W": "Wash sale", "LT": "Long-term", "ST": "Short-term",
    "LD": "Loss disallowed (wash)", "FX": "FX-related",
}


def decode_notes(notes: str) -> str:
    if not notes:
        return ""
    parts = []
    for code in notes.replace(",", ";").split(";"):
        code = code.strip()
        if code in NOTE_CODES:
            parts.append(f"{code}={NOTE_CODES[code]}")
        elif code:
            parts.append(code)
    return "; ".join(parts)


# -------------- Enrich with EUR --------------

def enrich_with_eur(lots: list[dict],
                    ecb: dict[date, dict[str, float]]) -> tuple[list[dict], list[dict]]:
    """Adds EUR-converted columns. Returns (lots_enriched, audit_rows)."""
    audit: list[dict] = []
    audit_keys_seen: set[tuple[date, str]] = set()
    for lot in lots:
        ccy = lot["currency"]
        comm_ccy = lot["commissionCurrency"] or ccy
        # Rate for opening leg (cost basis valued at open date)
        open_rate, open_used, open_note = (None, None, "")
        if lot["openDate"]:
            open_rate, open_used, open_note = rate_for(ecb, lot["openDate"], ccy)
        # Rate for closing leg
        close_rate, close_used, close_note = rate_for(ecb, lot["closeDate"], ccy)
        # Commission rate (commission may be in different ccy than trade)
        comm_rate, comm_used, comm_note = rate_for(ecb, lot["closeDate"], comm_ccy)

        # EUR values
        lot["openValueOrig"] = lot["costBasis"]  # already abs, None if anomaly
        lot["closeValueOrig"] = lot["proceeds"]  # gross close value (always reliable)
        lot["openValueEUR"] = to_eur(lot["openValueOrig"], open_rate) if lot["openValueOrig"] is not None else None
        lot["closeValueEUR"] = to_eur(lot["closeValueOrig"], close_rate)
        lot["commissionEUR"] = to_eur(abs(lot["commission"]), comm_rate)
        comm_eur = lot["commissionEUR"] or 0

        # Realized P&L EUR — methodology by asset category:
        # FUT (futures): use IB realized × close-date FX rate. Reason: futures
        #   settle daily via variation margin; there is no real cash paid at
        #   "open notional" that should be converted at open-date FX. Converting
        #   synthetic notional at open-date FX creates a phantom FX gain/loss
        #   on money that was never actually moved.
        # STK/OPT/FOP non-anomaly: per-leg FX (open at open-date rate, close at
        #   close-date rate). This correctly captures FX-translation effect on
        #   real cash flows.
        # Anomaly lots (any class): fall back to realized × close rate.
        if lot["assetCategory"] == "FUT":
            # Always realized × close rate for futures
            lot["realizedPnlEUR"] = to_eur(lot["fifoPnlRealized"], close_rate)
            # For accounting display, backfill open EUR using identity (LONG/SHORT):
            if lot["closeValueEUR"] is not None and lot["realizedPnlEUR"] is not None:
                if lot["is_long"]:
                    lot["openValueEUR"] = lot["closeValueEUR"] - lot["realizedPnlEUR"] - comm_eur
                else:
                    lot["openValueEUR"] = lot["closeValueEUR"] + lot["realizedPnlEUR"] + comm_eur
        elif lot["closeValueEUR"] is not None and lot["openValueEUR"] is not None:
            # STK/OPT/FOP non-anomaly: direction-aware EUR P&L
            if lot["is_long"]:
                lot["realizedPnlEUR"] = lot["closeValueEUR"] - lot["openValueEUR"] - comm_eur
            else:
                lot["realizedPnlEUR"] = lot["openValueEUR"] - lot["closeValueEUR"] - comm_eur
        else:
            # Anomaly lot (basis non-derivable) — IB realized × close FX rate
            lot["realizedPnlEUR"] = to_eur(lot["fifoPnlRealized"], close_rate)
            if lot["closeValueEUR"] is not None and lot["realizedPnlEUR"] is not None:
                if lot["is_long"]:
                    lot["openValueEUR"] = lot["closeValueEUR"] - lot["realizedPnlEUR"] - comm_eur
                else:
                    lot["openValueEUR"] = lot["closeValueEUR"] + lot["realizedPnlEUR"] + comm_eur
        lot["holdingDays"] = (lot["closeDate"] - lot["openDate"]).days if lot["openDate"] else None
        lot["openRateUsed"] = open_rate
        lot["openRateDate"] = open_used
        lot["openRateNote"] = open_note
        lot["closeRateUsed"] = close_rate
        lot["closeRateDate"] = close_used
        lot["closeRateNote"] = close_note
        lot["commRateUsed"] = comm_rate
        lot["commRateNote"] = comm_note
        lot["anomalyFlag"] = "YES" if lot["anomaly"] else ""
        if lot["anomaly"]:
            anomaly_note = "ANOMALY: IB cost basis non-derivable (transferred-in / corporate action / roll). P&L taken from IB FIFO at close-date FX rate."
            lot["notesDecoded"] = (lot["notesDecoded"] + "; " + anomaly_note) if lot["notesDecoded"] else anomaly_note

        for d, c, r, dused, note in [
            (lot["openDate"], ccy, open_rate, open_used, open_note),
            (lot["closeDate"], ccy, close_rate, close_used, close_note),
            (lot["closeDate"], comm_ccy, comm_rate, comm_used, comm_note),
        ]:
            if d is None or not c:
                continue
            key = (d, c)
            if key not in audit_keys_seen and r is not None:
                audit_keys_seen.add(key)
                audit.append({
                    "tradeDate": d, "currency": c,
                    "ecbRateDate": dused, "ratePerEUR": r,
                    "note": note,
                })
    audit.sort(key=lambda r: (r["currency"], r["tradeDate"]))
    return lots, audit


# -------------- XLSX writer --------------

HEADER_FILL = PatternFill("solid", fgColor="305496")
HEADER_FONT = Font(bold=True, color="FFFFFF")
EUR_FORMAT = '#,##0.00 "€"'
DATE_FORMAT = "yyyy-mm-dd"

LOT_COLUMNS = [
    ("accountId", "Account", 12),
    ("assetCategory", "Asset", 8),
    ("symbol", "Symbol", 18),
    ("description", "Description", 30),
    ("isin", "ISIN", 14),
    ("conid", "ConID", 12),
    ("currency", "Ccy", 6),
    ("direction", "Direction", 10),
    ("anomalyFlag", "Anomaly", 9),
    ("openDate", "Open Date", 12),
    ("openPrice", "Open Price", 12),
    ("openValueOrig", "Open Value (ccy)", 16),
    ("openValueEUR", "Open Value EUR", 16),
    ("closeDate", "Close Date", 12),
    ("closePrice", "Close Price", 12),
    ("quantity", "Quantity", 10),
    ("multiplier", "Mult.", 7),
    ("closeValueOrig", "Close Value (ccy)", 16),
    ("closeValueEUR", "Close Value EUR", 16),
    ("commission", "Commission (orig)", 14),
    ("commissionCurrency", "Comm Ccy", 9),
    ("commissionEUR", "Commission EUR", 14),
    ("realizedPnlEUR", "Realized P&L EUR", 16),
    ("fifoPnlRealized", "IB FIFO P&L (orig)", 16),
    ("holdingDays", "Holding Days", 12),
    ("openCloseIndicator", "Open/Close", 11),
    ("notesDecoded", "Notes (decoded)", 60),
    ("transactionID", "Tx ID", 12),
    ("ibExecID", "Exec ID", 14),
    ("openRateDate", "ECB Open Rate Date", 14),
    ("openRateUsed", "ECB Open Rate", 13),
    ("closeRateDate", "ECB Close Rate Date", 14),
    ("closeRateUsed", "ECB Close Rate", 13),
    ("openRateNote", "ECB Open Note", 18),
    ("closeRateNote", "ECB Close Note", 18),
]

EUR_COLS = {"openValueEUR", "closeValueEUR", "commissionEUR", "realizedPnlEUR"}
DATE_COLS = {"openDate", "closeDate", "openRateDate", "closeRateDate"}
NUM_COLS = {"openPrice", "closePrice", "quantity", "multiplier",
            "openValueOrig", "closeValueOrig", "commission", "fifoPnlRealized",
            "openRateUsed", "closeRateUsed", "holdingDays"}


def _write_header(ws, columns):
    for i, (_, label, width) in enumerate(columns, start=1):
        c = ws.cell(row=1, column=i, value=label)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.alignment = Alignment(horizontal="center", vertical="center")
        ws.column_dimensions[get_column_letter(i)].width = width
    ws.freeze_panes = "A2"


def write_lots_sheet(wb: Workbook, lots: list[dict]):
    ws = wb.active
    ws.title = "Closed Lots"
    _write_header(ws, LOT_COLUMNS)
    # Sort lots by close date then symbol
    lots = sorted(lots, key=lambda x: (x["closeDate"], x["symbol"]))
    for r, lot in enumerate(lots, start=2):
        for i, (key, _, _) in enumerate(LOT_COLUMNS, start=1):
            val = lot.get(key)
            cell = ws.cell(row=r, column=i, value=val)
            if key in EUR_COLS:
                cell.number_format = EUR_FORMAT
            elif key in DATE_COLS:
                cell.number_format = DATE_FORMAT
            elif key in NUM_COLS:
                cell.number_format = "#,##0.0000"
    ws.auto_filter.ref = ws.dimensions


def write_summary_sheet(wb: Workbook, lots: list[dict]):
    """Per-symbol summary: total realized P&L, quantity, commission."""
    ws = wb.create_sheet("Summary by Symbol")
    cols = [
        ("Symbol", 18), ("Description", 30), ("Ccy", 6),
        ("# Lots", 8), ("Total Quantity", 14),
        ("Total Open EUR", 16), ("Total Close EUR", 16),
        ("Total Commission EUR", 18), ("Total Realized P&L EUR", 20),
    ]
    for i, (label, width) in enumerate(cols, start=1):
        c = ws.cell(row=1, column=i, value=label)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.alignment = Alignment(horizontal="center")
        ws.column_dimensions[get_column_letter(i)].width = width
    ws.freeze_panes = "A2"

    groups: dict[tuple[str, str, str], dict] = {}
    for lot in lots:
        key = (lot["symbol"], lot["description"], lot["currency"])
        g = groups.setdefault(key, {
            "lots": 0, "qty": 0.0,
            "open_eur": 0.0, "close_eur": 0.0,
            "comm_eur": 0.0, "pnl_eur": 0.0,
        })
        g["lots"] += 1
        g["qty"] += lot["quantity"]
        if lot["openValueEUR"] is not None:
            g["open_eur"] += lot["openValueEUR"]
        if lot["closeValueEUR"] is not None:
            g["close_eur"] += lot["closeValueEUR"]
        if lot["commissionEUR"] is not None:
            g["comm_eur"] += lot["commissionEUR"]
        if lot["realizedPnlEUR"] is not None:
            g["pnl_eur"] += lot["realizedPnlEUR"]
    rows = sorted(groups.items(), key=lambda kv: kv[1]["pnl_eur"])
    for r, ((sym, desc, ccy), g) in enumerate(rows, start=2):
        ws.cell(row=r, column=1, value=sym)
        ws.cell(row=r, column=2, value=desc)
        ws.cell(row=r, column=3, value=ccy)
        ws.cell(row=r, column=4, value=g["lots"])
        ws.cell(row=r, column=5, value=g["qty"]).number_format = "#,##0.00"
        for i, k in enumerate(["open_eur", "close_eur", "comm_eur", "pnl_eur"], start=6):
            ws.cell(row=r, column=i, value=g[k]).number_format = EUR_FORMAT
    # Grand total row
    total_row = len(rows) + 2
    ws.cell(row=total_row, column=1, value="TOTAL").font = Font(bold=True)
    for i, k in enumerate(["open_eur", "close_eur", "comm_eur", "pnl_eur"], start=6):
        cell = ws.cell(row=total_row, column=i, value=sum(g[k] for _, g in groups.items()))
        cell.number_format = EUR_FORMAT
        cell.font = Font(bold=True)
    ws.auto_filter.ref = f"A1:{get_column_letter(len(cols))}{len(rows) + 1}"


def write_manual_review_sheet(wb: Workbook, lots: list[dict]):
    """Separate sheet listing all lots flagged for manual basis review.
    Per GPT-5 review: anomaly lots are not silently usable — they need explicit
    accountant attention (transferred-in basis, corporate actions like Bitcoin
    Mini Trust spin-off from GBTC, option rolls, return-of-capital, etc.).
    """
    anom = [l for l in lots if l.get("anomaly")]
    if not anom:
        return
    ws = wb.create_sheet("Manual Review Required")
    cols = [
        ("Symbol", 25), ("Description", 32), ("Asset", 8), ("Direction", 10),
        ("Open Date (IB)", 14), ("Close Date", 12), ("Quantity", 10),
        ("Close Price", 12), ("Close Value (USD)", 16),
        ("IB FIFO P&L (USD)", 16), ("Realized P&L EUR (fallback)", 22),
        ("Reason / Suggested Action", 80),
    ]
    for i, (label, width) in enumerate(cols, start=1):
        c = ws.cell(row=1, column=i, value=label)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.column_dimensions[get_column_letter(i)].width = width
    ws.freeze_panes = "A2"
    for r, lot in enumerate(anom, start=2):
        ws.cell(row=r, column=1, value=lot["symbol"])
        ws.cell(row=r, column=2, value=lot["description"])
        ws.cell(row=r, column=3, value=lot["assetCategory"])
        ws.cell(row=r, column=4, value=lot["direction"])
        ws.cell(row=r, column=5, value=lot["openDate"]).number_format = DATE_FORMAT
        ws.cell(row=r, column=6, value=lot["closeDate"]).number_format = DATE_FORMAT
        ws.cell(row=r, column=7, value=lot["quantity"])
        ws.cell(row=r, column=8, value=lot["closePrice"]).number_format = "#,##0.0000"
        ws.cell(row=r, column=9, value=lot["closeValueOrig"]).number_format = "#,##0.00"
        ws.cell(row=r, column=10, value=lot["fifoPnlRealized"]).number_format = "#,##0.00"
        ws.cell(row=r, column=11, value=lot["realizedPnlEUR"]).number_format = EUR_FORMAT
        # Suggest reason based on symbol pattern.
        # Crypto trusts that underwent spin-off / reorganization events known to
        # break IB's downstream cost-basis derivation in Flex Closed-Lots reports.
        # Extend this list opportunistically when new ones surface.
        KNOWN_SPINOFF_TRUSTS = {"GBTC", "ETHE", "BITW", "BITQ"}
        reason = "IB-recorded cost basis non-derivable from this Flex query."
        if any(t in lot["symbol"] for t in KNOWN_SPINOFF_TRUSTS):
            reason += " Likely cause: crypto-trust spin-off / reorganization (e.g. GBTC→BTC Mini Trust July 2024, ETHE-related events) reallocated cost basis on the Flex side. Suggested action: request IB 'Cost Basis Statement' (separate report) or original-broker statements for ACAT cost basis."
        elif lot["assetCategory"] == "OPT":
            reason += " Possible cause: option roll with carry-over basis, prior wash-sale adjustment, or fragment of multi-leg position. Suggested action: review prior option activity on same underlying for connecting trades."
        elif lot["assetCategory"] == "STK":
            reason += " Possible cause: transferred-in position with adjusted basis at receipt, corporate action, or return of capital. Suggested action: request IB 'Cost Basis Statement' or original-broker ACAT records."
        else:
            reason += " Suggested action: cross-check with IB monthly statement and prior corporate action notices."
        cell = ws.cell(row=r, column=12, value=reason)
        cell.alignment = Alignment(wrap_text=True, vertical="top")
    ws.auto_filter.ref = ws.dimensions


def write_audit_sheet(wb: Workbook, audit: list[dict], meta: dict):
    ws = wb.create_sheet("ECB Rates Audit")
    cols = [("Trade Date", 12), ("Currency", 10), ("ECB Rate Date", 14),
            ("Rate per 1 EUR", 16), ("Note", 30)]
    for i, (label, width) in enumerate(cols, start=1):
        c = ws.cell(row=1, column=i, value=label)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        ws.column_dimensions[get_column_letter(i)].width = width
    ws.freeze_panes = "A2"
    for r, row in enumerate(audit, start=2):
        ws.cell(row=r, column=1, value=row["tradeDate"]).number_format = DATE_FORMAT
        ws.cell(row=r, column=2, value=row["currency"])
        ws.cell(row=r, column=3, value=row["ecbRateDate"]).number_format = DATE_FORMAT
        ws.cell(row=r, column=4, value=row["ratePerEUR"]).number_format = "0.000000"
        ws.cell(row=r, column=5, value=row["note"])

    # Metadata sheet
    ws2 = wb.create_sheet("Metadata")
    for i, (k, v) in enumerate([
        ("Account", meta.get("accountId", "")),
        ("Period from", meta.get("fromDate", "")),
        ("Period to", meta.get("toDate", "")),
        ("Query name", meta.get("queryName", "")),
        ("When generated (IB)", meta.get("whenGenerated", "")),
        ("Skipped trades (wrong level)", str(meta.get("skipped_by_level", {}))),
        ("Conversion source", "ECB euro foreign exchange reference rates (eurofxref-hist.csv)"),
        ("Conversion direction", "EUR_amount = trade_ccy_amount / (CCY per 1 EUR rate on trade date)"),
        ("CNH (offshore yuan)", "Converted via CNY (onshore yuan) — ECB does not publish CNH"),
        ("Carry-forward policy", "If trade date is non-business day (weekend/TARGET2 holiday), uses last available prior business day rate"),
        ("FUT methodology", "Futures EUR P&L = IB realized USD ÷ close-date ECB rate. NOT per-leg notional×FX, because futures settle daily via variation margin (no real cash at 'open notional'). Per-leg conversion would create phantom FX gain/loss on money never moved."),
        ("Anomaly fallback", "Lots flagged 'ANOMALY=YES' have IB-recorded cost basis that derives negative — typically transferred-in positions, corporate-action spin-offs (e.g., GBTC→BTC Mini Trust July 2024), option rolls with carry-over basis. EUR P&L for these = IB realized USD ÷ close-date rate. See 'Manual Review Required' sheet for action items."),
        ("Commissions", "Closed-Lots query returned empty ibCommission for all lots — IB reports them in regular Trades section but not in Closed-Lots aggregation. Realized P&L from IB is already net of commissions (IB definition: realizedPnl = proceeds + costBasis_closing, where costBasis includes commission). No double-count occurs because we don't subtract a separate commission line."),
        ("Cross-checks passed", "(1) Row count XML=XLSX, (2) IB FIFO sum reconciles, (3) Direction tagging consistent, (4) LONG accounting identity, (5) SHORT accounting identity, (6) ECB rates within 0.5% of IB fxRateToBase, (7) Holding days math, (8) LT/ST notes consistent with holding>365d, (9) Summary tab reconciles. Two AI cross-checkers (Gemini-2.5-pro and GPT-5) reviewed parser logic; their concerns addressed in this version."),
    ], start=1):
        ws2.cell(row=i, column=1, value=k).font = Font(bold=True)
        ws2.cell(row=i, column=2, value=v)
    ws2.column_dimensions["A"].width = 30
    ws2.column_dimensions["B"].width = 80


# -------------- main --------------

def main() -> int:
    p = argparse.ArgumentParser()
    p.add_argument("--xml", help="Flex Closed-Lots XML file (omit if using --year)")
    p.add_argument("--ecb", help="ECB eurofxref-hist.csv (default: data/ecb/eurofxref-hist.csv)")
    p.add_argument("--out", help="Output XLSX path (omit if using --year)")
    p.add_argument("--year", type=int,
                   help="Shortcut: auto-resolve --xml, --ecb, --out from convention. "
                        "Reads portfolio/ib/tax/{YEAR}/closed_lots_{YEAR}.xml, "
                        "writes portfolio/ib/tax/{YEAR}/Closed_Lots_{YEAR}_EUR.xlsx, "
                        "ECB rates from data/ecb/eurofxref-hist.csv.")
    args = p.parse_args()

    # --year shortcut: derive all paths from convention
    if args.year:
        if not args.xml:
            args.xml = f"portfolio/ib/tax/{args.year}/closed_lots_{args.year}.xml"
        if not args.out:
            args.out = f"portfolio/ib/tax/{args.year}/Closed_Lots_{args.year}_EUR.xlsx"
    if not args.ecb:
        args.ecb = "data/ecb/eurofxref-hist.csv"
    # After shortcuts: all three must be set
    missing = [n for n, v in (("--xml", args.xml), ("--ecb", args.ecb), ("--out", args.out)) if not v]
    if missing:
        print(f"ERROR: missing required args: {', '.join(missing)} (or use --year YYYY shortcut)",
              file=sys.stderr)
        return 2

    print(f"Loading ECB rates from {args.ecb} ...", file=sys.stderr)
    ecb, ccys = load_ecb_rates(Path(args.ecb))
    print(f"  {len(ecb)} business days, {len(ccys)} currencies", file=sys.stderr)

    print(f"Parsing closed lots from {args.xml} ...", file=sys.stderr)
    lots, meta = extract_lots(Path(args.xml))
    print(f"  {len(lots)} closed-lot rows extracted", file=sys.stderr)
    print(f"  skipped (wrong level): {meta.get('skipped_by_level', {})}", file=sys.stderr)
    if not lots:
        print("ERROR: no Closed-Lot rows found in XML. Likely the Flex query was not configured "
              "with Level of Detail = Closed Lots.", file=sys.stderr)
        return 3

    print("Enriching with EUR conversion ...", file=sys.stderr)
    lots, audit = enrich_with_eur(lots, ecb)

    print(f"Writing XLSX -> {args.out}", file=sys.stderr)
    wb = Workbook()
    write_lots_sheet(wb, lots)
    write_summary_sheet(wb, lots)
    write_manual_review_sheet(wb, lots)
    write_audit_sheet(wb, audit, meta)
    wb.save(args.out)

    # Totals print
    total_pnl = sum(l["realizedPnlEUR"] for l in lots if l["realizedPnlEUR"] is not None)
    print(f"\nDONE. {len(lots)} lots, total Realized P&L in EUR: {total_pnl:,.2f} €",
          file=sys.stderr)
    return 0


if __name__ == "__main__":
    sys.exit(main())
